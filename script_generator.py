import streamlit as st
import pandas as pd
import openpyxl
import os
from connector import get_teradata_connection, get_saved_connections

# Base mappings
teradata_to_azure = {
    'CF': 'VARCHAR',
    'CV': 'VARCHAR',
    'D': 'DECIMAL(18,4)',
    'I': 'INT',
    'DA': 'DATE',
    'TS': 'DATETIME',
    'I8': 'BIGINT',
    'I1': 'TINYINT'
}

azure_to_teradata = {v: k for k, v in teradata_to_azure.items()}

teradata_to_databricks = {
    'CF': 'STRING',
    'CV': 'STRING',
    'D': 'DECIMAL(18,4)',
    'I': 'INT',
    'DA': 'DATE',
    'TS': 'TIMESTAMP',
    'I8': 'BIGINT',
    'I1': 'BYTE'
}

databricks_to_teradata = {v: k for k, v in teradata_to_databricks.items()}

azure_to_databricks = {
    'VARCHAR': 'STRING',
    'DECIMAL(18,4)': 'DECIMAL(18,4)',
    'INT': 'INT',
    'DATE': 'DATE',
    'DATETIME': 'TIMESTAMP',
    'BIGINT': 'BIGINT',
    'TINYINT': 'BYTE'
}

databricks_to_azure = {v: k for k, v in azure_to_databricks.items()}

def get_mapping(source, target):
    if source == "Teradata" and target == "Azure SQL DB":
        return teradata_to_azure
    elif source == "Azure SQL DB" and target == "Teradata":
        return azure_to_teradata
    elif source == "Teradata" and target == "Databricks":
        return teradata_to_databricks
    elif source == "Databricks" and target == "Teradata":
        return databricks_to_teradata
    elif source == "Azure SQL DB" and target == "Databricks":
        return azure_to_databricks
    elif source == "Databricks" and target == "Azure SQL DB":
        return databricks_to_azure
    else:
        return {}

def map_column_type(column_type, mapping_dict):
    return mapping_dict.get(column_type.strip(), column_type)

def generate_ddl_script(metadata_df, target_schema, target_table):
    ddl_lines = [f"CREATE TABLE {target_schema}.{target_table} ("]
    for idx, row in metadata_df.iterrows():
        col_name = row['ColumnName']
        col_type = row['Mapped DataType']
        nullable = "NULL" if row['Nullable'] == 'Y' else "NOT NULL"
        ddl_lines.append(f"  {col_name} {col_type} {nullable},")
    ddl_lines[-1] = ddl_lines[-1].rstrip(',')
    ddl_lines.append(");")
    return "\n".join(ddl_lines)

def generate_etl_script(metadata_df, source_schema, source_table, target_schema, target_table):
    columns = metadata_df['ColumnName'].tolist()
    column_list = ", ".join(columns)
    return f"INSERT INTO {target_schema}.{target_table} ({column_list})\nSELECT {column_list} FROM {source_schema}.{source_table};"

def run_script_generator_ui():
    st.title("Document Generator")
    st.markdown("Generate STTM, DDL, and ETL/ELT scripts from metadata with platform-specific data types.")

    connector_types = ["Teradata", "Azure SQL DB", "Databricks"]
    source_db = st.selectbox("Select Source Database", connector_types)
    target_db = st.selectbox("Select Target Database", connector_types)

    source_key_map = {
        "Teradata": "teradata",
        "Azure SQL DB": "azuresql",
        "Databricks": "databricks"
    }
    source_key = source_key_map.get(source_db, source_db.lower().replace(" ", ""))

    saved_conns = get_saved_connections(source_key)
    if not saved_conns:
        st.warning(f"No saved {source_db} connections found. Please create one in the Connector panel from Home page.")
        st.stop()
    selected_conn = st.selectbox(f"Select {source_db} Connection", saved_conns)

    schema_name = st.text_input("Source Schema Name")
    table_name = st.text_input("Source Table Name")
    target_schema = st.text_input("Target Schema Name")
    target_table = st.text_input("Target Table Name")
    # Template selection
    st.markdown("### Choose STTM Template")
    template_option = st.radio("Select template source:", ["Use default template", "Upload your own template"])

    if template_option == "Upload your own template":
        template_file = st.file_uploader("Upload STTM Template (.xlsx)", type=["xlsx"])
        if not template_file:
            st.warning("Please upload a template to proceed.")
            st.stop()
    else:
        template_file = "sttm_template.xlsx"  # Path to default template in the same folder


    if st.button("Fetch Metadata and Generate Scripts"):
        if not schema_name or not table_name or not target_schema or not target_table or not template_file:
            st.error("Please fill in all required fields and upload the template.")
        else:
            try:
                if source_db == "Teradata":
                    connection = get_teradata_connection(selected_conn)
                    if connection is None:
                        st.stop()
                    query = f"""
                    SELECT
                        ColumnName,
                        ColumnType,
                        ColumnLength,
                        Nullable
                    FROM DBC.ColumnsV
                    WHERE DatabaseName = '{schema_name}' AND TableName = '{table_name}'
                    ORDER BY ColumnId;
                    """
                    metadata_df = pd.read_sql(query, connection)
                    connection.close()
                else:
                    st.error(f"Metadata fetching for {source_db} is not yet implemented, will be added soon :).")
                    st.stop()

                mapping_dict = get_mapping(source_db, target_db)
                metadata_df['Mapped DataType'] = metadata_df['ColumnType'].apply(lambda x: map_column_type(x, mapping_dict))
                st.success("Metadata fetched and mapped successfully!")
                st.dataframe(metadata_df)

                wb = openpyxl.load_workbook(template_file)
                sheet = wb.active
                for idx, row in metadata_df.iterrows():
                    sheet.cell(row=idx + 3, column=1, value=schema_name)
                    sheet.cell(row=idx + 3, column=2, value=table_name)
                    sheet.cell(row=idx + 3, column=3, value=row['ColumnName'])
                    sheet.cell(row=idx + 3, column=4, value=row['ColumnType'])
                    sheet.cell(row=idx + 3, column=5, value=row['Nullable'])
                    sheet.cell(row=idx + 3, column=6, value="Direct Mapping")
                    sheet.cell(row=idx + 3, column=7, value=target_schema)
                    sheet.cell(row=idx + 3, column=8, value=target_table)
                    sheet.cell(row=idx + 3, column=9, value=row['ColumnName'])
                    sheet.cell(row=idx + 3, column=10, value=row['Mapped DataType'])
                    sheet.cell(row=idx + 3, column=11, value=row['Nullable'])
                    sheet.cell(row=idx + 3, column=12, value="")  # PrimaryKey placeholder

                output_dir = "output"
                os.makedirs(output_dir, exist_ok=True)
                sttm_path = os.path.join(output_dir, f"{table_name}_sttm.xlsx")
                wb.save(sttm_path)
                st.success("STTM Excel generated.")
                with open(sttm_path, "rb") as f:
                    st.download_button("Download STTM Excel", f, file_name=f"{table_name}_sttm.xlsx")

                ddl_script = generate_ddl_script(metadata_df, target_schema, target_table)
                st.download_button("Download DDL Script", ddl_script, file_name=f"{target_table}_ddl.sql")

                etl_script = generate_etl_script(metadata_df, schema_name, table_name, target_schema, target_table)
                st.download_button("Download ETL Script", etl_script, file_name=f"{target_table}_etl.sql")

            except Exception as e:
                st.error(f"Error: {e}")
